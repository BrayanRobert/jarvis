from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from time import sleep
from datetime import date
from datetime import timedelta
import openpyxl

options = Options()
options.headless = False  # Executar de forma visível ou oculta
navegador = webdriver.Chrome(options=options)

# Células do excel
vetor = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']

# Selecionando as unidades: 1- Nova Iguaçu, 2- Nilópolis, 3- Miguel Couto, 4- São João de Meriti
uni = 1

# Selecionando o procedimento:
# 1 - US Geral
# 2 - US Vascular
# 3 - US Obstétrico

procedimentos = [1, 2, 3]

# Função responsável por conectar com a página de Login
def Connection(navegador):
    link = "https://app.feegow.com/main/?P=Login&qs="

    navegador.get(url=link)
    sleep(1)
# Função responsável por logar o usuário
def Login(navegador,user,password):

    # Colocando o email do usuário
    inputUsuario = navegador.find_element(by=By.ID, value="User")
    inputUsuario.send_keys(user)
    sleep(1)

    # Colocando a senha do usuário
    inputSenha = navegador.find_element(by=By.ID, value="password")
    inputSenha.send_keys(password)

    # Clicando no botão para logar
    buttonLogin = navegador.find_element(by=By.ID, value="Entrar")
    buttonLogin.click()
    sleep(1)
# Função responsável para definir os requisitos (Abrir agenda, escolher a múltipla e abrir o campo dos locais)
def Requirements(navegador):
    # Clicando na opção AGENDA
    agenda = navegador.find_element(by=By.XPATH, value='/html/body/aside/header/ul[1]/li[1]/a')
    agenda.click()
    sleep(1)

    # Escolhendo a opção AGENDA MÚLTIPLA
    multipla = navegador.find_element(by=By.XPATH, value='/html/body/aside/header/ul[1]/li[1]/ul/li[4]/a')
    multipla.click()
    sleep(0.5)

    # Clicando no campo de LOCAIS
    campoLocal = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/button')
    campoLocal.click()

    return navegador
# Função responsável para escolher as unidades
def Units(navegador,unidades):

    match unidades:
        case 1:
        # Escolhendo os LOCAIS - Nova Iguaçu
            local1 = navegador.find_element(by=By.XPATH,value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[3]/a/label')
            local1.click()
            sleep(0.5)

            local2 = navegador.find_element(by=By.XPATH,value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[12]/a/label')
            local2.click()
            sleep(0.5)

            local3 = navegador.find_element(by=By.XPATH,value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[9]/a/label')
            local3.click()
            sleep(0.5)

            local4 = navegador.find_element(by=By.XPATH,value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[10]/a/label')
            local4.click()
            sleep(0.5)

            local5 = navegador.find_element(by=By.XPATH,value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[11]/a/label')
            local5.click()
            sleep(0.5)

        case 2:
            # Escolhendo o LOCAL - Nilópolis
            local = navegador.find_element(by=By.XPATH,value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[14]/a/label')
            local.click()
            sleep(0.5)

        case 3:
             # Escolhendo o LOCAL - Miguel Couto
             local = navegador.find_element(by=By.XPATH,value='//*[@id="qflocais"]/div/ul/li[13]/a/label')
             local.click()
             sleep(0.5)

        case 4:
            # Escolhendo o LOCAL - São João
            local = navegador.find_element(by=By.XPATH,value='//*[@id="qflocais"]/div/ul/li[15]/a/label')
            local.click()
            sleep(0.5)
# Função que seleciona a lista de MÉDICOS mediante ao GRUPO DE PROCEDIMENTO
def Group():
    match procedimentos:
        case 1:
            lista = [
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[81]/a/label',
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[119]/a/label',
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[82]/a/label',
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[89]/a/label',
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[77]/a/label',
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[20]/a/label',
                '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[64]/a/label'
            ]
            cel = [31, 32, 33, 34, 35, 36, 37]
        case 2:
            lista = ['/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/ul/li[81]/a/label']
            cel = [40]
        case 3:
            lista = [

            ]
            cel = []
        case 4:
            lista = [

            ]
            cel = []
        case 5:
            lista = [

            ]
            cel = []
        case 6:
            lista = [

            ]
            cel = []

    return lista, cel
# Função responsável para coletar os dados da página (WEBSCRAPING)
def WebScraping(navegador):
    for proc in list(procedimentos):
        procedimentos[proc]
        lista = Group()[0]
        cel = Group()[1]
        for i in range(len(lista)):
            aux1 = lista[i]
            if aux1 == lista[0]:
                # Clicando no campo de PROFISSIONAIS
                campoProf = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/button')
                campoProf.click()
                sleep(1)

                # Selecionando o PROFISSIONAL
                selProf = navegador.find_element(by=By.XPATH, value=aux1)
                selProf.click()
                sleep(1)
            else:
                # Clicando no campo de PROFISSIONAIS
                campoProf = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[2]/div/button')
                campoProf.click()
                sleep(1)

                # Retirando o profissional selecionado anteriormente
                selProf = navegador.find_element(by=By.XPATH, value=lista[i - 1])
                selProf.click()
                sleep(1)

                # Selecionando o PROFISSIONAL
                selProf = navegador.find_element(by=By.XPATH, value=aux1)
                selProf.click()
                sleep(1)




            for j in range(23):
                aux = date.today() + timedelta(days=j + 1)
                data = aux.strftime('%d/%m/%Y')

                print(f'A data é: {data}')

                if data == '01/08/2022':

                    # Clicando para passar para a próxima agenda no calendário, para o selenium conseguir encontrar o elemento certo
                    proximo = navegador.find_element(by=By.XPATH,value='//*[@id="tblCalendario"]/thead/tr[1]/th[3]')
                    proximo.click()
                    sleep(1)

                    # Encontrar o campo da data e clicar nele
                    data = navegador.find_element(by=By.ID, value=data)
                    data.click()
                    sleep(2)

                    # Conta o número de vagas e imprime o resultado
                    vagas = navegador.find_elements(by=By.CSS_SELECTOR, value='.btn-info')
                    resultado = len(vagas)
                    print(resultado)
                    sleep(2)

                    # Atribuindo os valores as células do arquivo Excel desejado
                    book = openpyxl.load_workbook("procedimentos.xlsx")
                    ws = book.worksheets[0]
                    ws[vetor[j]+str(cel[i])].value = resultado
                    book.save("procedimentos.xlsx")
                    print(f"Vagas Disponíveis: {resultado} salvo na célula {vetor[j] + str(cel[i])}!")

                else:
                    if data == '10/08/2022':
                        # Encontrar o campo da data e clicar nele
                        data = navegador.find_element(by=By.ID, value=data)
                        data.click()
                        sleep(2)

                        # Conta o número de vagas e imprime o resultado
                        vagas = navegador.find_elements(by=By.CSS_SELECTOR, value='.btn-info')
                        resultado = len(vagas)
                        print(resultado)
                        sleep(1)

                        # Atribuindo os valores as células do arquivo Excel desejado
                        book = openpyxl.load_workbook("procedimentos.xlsx")
                        ws = book.worksheets[0]
                        ws[vetor[j] + str(cel[i])].value = resultado
                        book.save("procedimentos.xlsx")
                        print(f"Vagas Disponíveis: {resultado} salvo na célula {vetor[j] + str(cel[i])}!")

                        # Clicando para voltar para a agenda anterior no calendário, para o selenium conseguir encontrar o elemento certo
                        anterior = navegador.find_element(by=By.XPATH, value='/html/body/aside/aside[1]/div[1]/ul/li[2]/div[3]/table/thead/tr[1]/th[1]')
                        anterior.click()
                        sleep(3)

                    else:
                        # Encontrar o campo da data e clicar nele
                        data = navegador.find_element(by=By.ID, value=data)
                        data.click()
                        sleep(2)

                        # Conta o número de vagas e imprime o resultado
                        vagas = navegador.find_elements(by=By.CSS_SELECTOR, value='.btn-info')
                        resultado = len(vagas)
                        print(resultado)
                        sleep(1)

                        # Atribuindo os valores as células do arquivo Excel desejado
                        book = openpyxl.load_workbook("procedimentos.xlsx")
                        ws = book.worksheets[0]
                        ws[vetor[j] + str(cel[i])].value = resultado
                        book.save("procedimentos.xlsx")
                        print(f"Vagas Disponíveis: {resultado} salvo na célula {vetor[j] + str(cel[i])}!")

    # Definindo o intervalo de datas que precisamos (D+1 até D+23)
    inicio = date.today() + timedelta(days=1)
    inicio_str = inicio.strftime('%d/%m/%Y')
    fim = date.today() + timedelta(days=23)
    fim_str = fim.strftime('%d/%m/%Y')

    # Fazendo conexão com a página de Login
    Connection(navegador=navegador)
    # Logando na página
    Login(navegador=navegador, user="brayansantos@segmedic.com.br",password="@Nadador01")
    # Fazendo os requisitos necessários
    Requirements(navegador=navegador)
    # Escolhendo as unidades
    Units(navegador=navegador, unidades=uni)
    # Fazendo o Webscraping
    WebScraping(navegador=navegador)

Connection(navegador=navegador)
Login(navegador=navegador, user='brayansantos@segmedic.com.br', password='@Nadador01')
Requirements(navegador=navegador)
Units(navegador=navegador, unidades=uni)
WebScraping(navegador=navegador)