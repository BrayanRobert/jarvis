from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from time import sleep
from datetime import date
from datetime import timedelta
import datetime
import openpyxl

options = Options()
options.headless = False  # Executar de forma visível ou oculta
navegador = webdriver.Chrome(options=options)

# Células do excel
cel = [16,17,18,19,23,24,25,26,27,28,29,30,31,32,33,34,35]
vetor = ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']

# Selecionando as unidades: 1- Nova Iguaçu, 2- Nilópolis, 3- Miguel Couto, 4- São João de Meriti
uni = [1]

# Lista de especialidades
lista = [
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[14]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[25]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[45]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[48]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[22]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[4]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[7]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[17]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[21]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[29]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[30]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[35]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[42]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[49]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[28]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[37]/a/label',
    '/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/ul/li[39]/a/label'
]

# Data inicial:

dataAtual = date.today()
nxt_mnth = dataAtual.replace(day=28) + datetime.timedelta(days=4)
dia = nxt_mnth.day
primeiroDia = nxt_mnth.strftime('%d/%m/%Y')


if dia == 2:
    nxt_mnth = dataAtual.replace(day=28) + datetime.timedelta(days=3)
    res = nxt_mnth.strftime('%d/%m/%Y')
    primeiroDia = res

else:
   primeiroDia = primeiroDia

# Função responsável por conectar com a página de Login
def Connection(navegador):
    link = "https://app.feegow.com/main/?P=Login&qs="

    navegador.get(url=link)
    sleep(1)
# Função responsável por logar o usuário
def Login(navegador,user,password):

    # Colocando o email do usuário
    inputUsuario = navegador.find_element(by=By.ID, value="User")
    inputUsuario.send_keys(user)
    sleep(1)

    # Colocando a senha do usuário
    inputSenha = navegador.find_element(by=By.ID, value="password")
    inputSenha.send_keys(password)

    # Clicando no botão para logar
    buttonLogin = navegador.find_element(by=By.ID, value="Entrar")
    buttonLogin.click()
    sleep(1)
# Função responsável para definir os requisitos (Abrir agenda, escolher a múltipla e abrir o campo dos locais)
def Requirements(navegador):
    # Clicando na opção AGENDA
    agenda = navegador.find_element(by=By.XPATH, value='/html/body/aside/header/ul[1]/li[1]/a')
    agenda.click()
    sleep(1)

    # Escolhendo a opção AGENDA MÚLTIPLA
    multipla = navegador.find_element(by=By.XPATH, value='/html/body/aside/header/ul[1]/li[1]/ul/li[4]/a')
    multipla.click()
    sleep(1)

    # Clicando no campo de LOCAIS
    campoLocal = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/button')
    campoLocal.click()

    return navegador
# Função responsável para escolher as unidades
def Units(navegador,unidades):

    match unidades:
        case 1:
        # Escolhendo os LOCAIS - Nova Iguaçu
            local1 = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[3]/a/label')
            local1.click()
            sleep(0.5)

            local2 = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[12]/a/label')
            local2.click()
            sleep(0.5)

            local3 = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[9]/a/label')
            local3.click()
            sleep(0.5)

            local4 = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[10]/a/label')
            local4.click()
            sleep(0.5)

            local5 = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[11]/a/label')
            local5.click()
            sleep(0.5)

        case 2:
            # Escolhendo o LOCAL - Nilópolis
            local = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[5]/div/ul/li[14]/a/label')
            local.click()
            sleep(0.5)

        case 3:
             # Escolhendo o LOCAL - Miguel Couto
             local = navegador.find_element(by=By.XPATH, value='//*[@id="qflocais"]/div/ul/li[13]/a/label')
             local.click()
             sleep(0.5)

        case 4:
            # Escolhendo o LOCAL - São João
            local = navegador.find_element(by=By.XPATH, value='//*[@id="qflocais"]/div/ul/li[15]/a/label')
            local.click()
            sleep(0.5)
# Função responsável para coletar os dados da página (WEBSCRAPING)
def WebScraping(navegador,lista):
    for i in range(len(lista)):

        aux1 = lista[i]

        if aux1 == lista[0]:
            print('Isso aconteceu')
            # Clicando no campo da especialidade
            buttonEspecialidade = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/button')
            buttonEspecialidade.click()
            sleep(2)

            # Seleciona a especialidade desejada
            sel_especialidade = navegador.find_element(by=By.XPATH, value=aux1)
            sel_especialidade.click()
            sleep(1)

            # Sobe a tela do navegador, para aparecer o botão de busca, e clicar posteriormente
            subirTela = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/header/div[1]/ol/li[3]')
            subirTela.click()
            sleep(1)

            # Busca a especialidade
            buscar = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[2]/div/button')
            buscar.click()
            sleep(1)
        else:
            # Clicando no campo da especialidade
            buttonEspecialidade = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/button')
            buttonEspecialidade.click()
            sleep(2)

            # Seleciona a especialidade desejada
            sel_especialidade = navegador.find_element(by=By.XPATH, value=lista[i - 1])
            sel_especialidade.click()
            sleep(1)

            # Sobe a tela do navegador, para aparecer o botão de busca, e clicar posteriormente
            subirTela = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/header/div[1]/ol/li[3]')
            subirTela.click()
            sleep(1)

            # Busca a especialidade
            buscar = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[2]/div/button')
            buscar.click()
            sleep(2)

            # Clicando no campo da especialidade
            buttonEspecialidade = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[1]/div[3]/div/button')
            buttonEspecialidade.click()
            sleep(2)

            # Seleciona a especialidade desejada
            sel_especialidade = navegador.find_element(by=By.XPATH, value=aux1)
            sel_especialidade.click()
            sleep(1)

            # Sobe a tela do navegador, para aparecer o botão de busca, e clicar posteriormente
            subirTela = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/header/div[1]/ol/li[3]')
            subirTela.click()
            sleep(1)

            # Busca a especialidade
            buscar = navegador.find_element(by=By.XPATH, value='/html/body/aside/section/section/aside/div/div[6]/div/form/div[2]/div/button')
            buscar.click()
            sleep(1)

        for j in range(23):
            aux = date.today() + timedelta(days=j + 1)
            data = aux.strftime('%d/%m/%Y')
            auxfim = date.today() + timedelta(days=23)
            datafim = auxfim.strftime('%d/%m/%Y')

            print(f'A data é: {data}')

            if data == primeiroDia:

                # Clicando para passar para a próxima agenda no calendário, para o selenium conseguir encontrar o elemento certo
                proximo = navegador.find_element(by=By.XPATH, value='//*[@id="tblCalendario"]/thead/tr[1]/th[3]')
                proximo.click()
                sleep(3)

                # Encontrar o campo da data e clicar nele
                data = navegador.find_element(by=By.ID, value=data)
                data.click()
                sleep(3)

                # Conta o número de vagas e imprime o resultado
                vagasNConf = navegador.find_elements(by=By.CSS_SELECTOR, value='.fa-question-circle.text-alert.badge-icon-status')
                vagasConf = navegador.find_elements(by=By.CSS_SELECTOR, value='.fa-grin.text-warning.badge-icon-status')
                vagas = len(vagasNConf) + len(vagasConf)
                resultado = vagas
                print(resultado)
                sleep(2)

                # Atribuindo os valores as células do arquivo Excel desejado
                book = openpyxl.load_workbook("VagasExcel.xlsx")
                ws = book.worksheets[1]
                ws[vetor[j]+str(cel[i])].value = resultado
                book.save("VagasExcel.xlsx")
                print(f"Vagas Disponíveis: {resultado} salvo na célula {vetor[j] + str(cel[i])}!")

            else:
                if data == datafim:
                    # Encontrar o campo da data e clicar nele
                    data = navegador.find_element(by=By.ID, value=data)
                    data.click()
                    sleep(2.5)

                    # Conta o número de vagas e imprime o resultado
                    vagasNConf = navegador.find_elements(by=By.CSS_SELECTOR, value='.fa-question-circle.text-alert.badge-icon-status')
                    vagasConf = navegador.find_elements(by=By.CSS_SELECTOR, value='.fa-grin.text-warning.badge-icon-status')
                    vagas = len(vagasNConf) + len(vagasConf)
                    resultado = vagas
                    print(resultado)
                    sleep(2)

                    # Atribuindo os valores as células do arquivo Excel desejado
                    book = openpyxl.load_workbook("VagasExcel.xlsx")
                    ws = book.worksheets[1]
                    ws[vetor[j] + str(cel[i])].value = resultado
                    book.save("VagasExcel.xlsx")
                    print(f"Vagas Disponíveis: {resultado} salvo na célula {vetor[j] + str(cel[i])}!")

                    # Clicando para voltar para a agenda anterior no calendário
                    anterior = navegador.find_element(by=By.XPATH, value='/html/body/aside/aside[1]/div[1]/ul/li[2]/div[3]/table/thead/tr[1]/th[1]')
                    anterior.click()
                    sleep(3)

                else:
                    # Encontrar o campo da data e clicar nele
                    data = navegador.find_element(by=By.ID, value=data)
                    data.click()
                    sleep(2.5)

                    # Conta o número de vagas e imprime o resultado
                    vagasNConf = navegador.find_elements(by=By.CSS_SELECTOR, value='.fa-question-circle.text-alert.badge-icon-status')
                    vagasConf = navegador.find_elements(by=By.CSS_SELECTOR, value='.fa-grin.text-warning.badge-icon-status')
                    vagas = len(vagasNConf) + len(vagasConf)
                    resultado = vagas
                    print(resultado)
                    sleep(2)

                    # Atribuindo os valores as células do arquivo Excel desejado
                    book = openpyxl.load_workbook("VagasExcel.xlsx")
                    ws = book.worksheets[1]
                    ws[vetor[j] + str(cel[i])].value = resultado
                    book.save("VagasExcel.xlsx")
                    print(f"Vagas Disponíveis: {resultado} salvo na célula {vetor[j] + str(cel[i])}!")

# Fazendo conexão com a página de Login
Connection(navegador=navegador)
# Logando na página
Login(navegador=navegador, user="brayansantos@segmedic.com.br", password="@Nadador01")
# Fazendo os requisitos necessários
Requirements(navegador=navegador)
# Escolhendo as unidades
Units(navegador=navegador, unidades=uni)
# Fazendo o Webscraping
WebScraping(navegador=navegador, lista=lista)